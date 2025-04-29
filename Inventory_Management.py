from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and select the active worksheet
wb = load_workbook('inventory.xlsx')
ws = wb.active

# Define red fill for low stock
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Loop through rows (starting from row 2 to skip header)
for row in ws.iter_rows(min_row=2, max_col=4):
    item = row[0].value
    qty = row[1].value
    price = row[2].value
    
    if qty is not None and price is not None:
        # Calculate total and set it
        row[3].value = qty * price

        # Highlight low stock items
        if qty < 5:
            for cell in row:
                cell.fill = red_fill

# Save the workbook
wb.save('inventory_updated.xlsx')